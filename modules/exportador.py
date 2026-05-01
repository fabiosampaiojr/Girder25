import os
import math
from datetime import datetime
from typing import List, Optional, Union, Dict, Any

def exportar_tabela(
    matriz: List[List],
    titulo: str,
    caminho_arquivo: str,
    cabecalho: Optional[List] = None,
    opcoes_exportacao: Optional[Dict] = None
) -> str:
    """
    Exporta uma tabela/matriz para diferentes formatos (Excel, PDF ou TXT).
    
    Parâmetros:
    -----------
    matriz : List[List]
        Lista de listas contendo os dados da tabela.
    titulo : str
        Título da tabela/documento.
    caminho_arquivo : str
        Caminho completo do arquivo a ser salvo (incluindo extensão: .xlsx, .pdf ou .txt)
    cabecalho : Optional[List]
        Lista com os nomes das colunas. Se None, a primeira linha da matriz
        será considerada como cabeçalho.
    opcoes_exportacao : Optional[Dict]
        Dicionário com opções adicionais para exportação.
        
    Retorna:
    --------
    str: Caminho do arquivo gerado.
    
    Exceções:
    ---------
    ValueError: Se o formato não for suportado.
    Exception: Para outros erros durante a exportação.
    """
    
    # Mapeamento de extensões para formatos internos
    MAPEAMENTO_EXTENSOES = {
        '.xlsx': 'xlsx',
        '.xls': 'xls',  # Suporte a formato mais antigo
        '.pdf': 'pdf',
        '.txt': 'txt'
    }
    
    def _formatar_valor(valor: Any) -> Union[str, int, float]:
        """
        Formata um valor para exibição na tabela.
        
        Regras de formatação:
        - Strings: mantém como estão
        - Inteiros: mantém como estão
        - Floats: arredonda para até 3 casas decimais, removendo zeros à direita
        
        Exemplos:
        - 1.000 -> 1
        - 2.45 -> 2.45
        - 3.001 -> 3.001
        - 4.000001 -> 4 (arredondado)
        - 5.123456 -> 5.123 (arredondado para 3 casas)
        
        Parâmetros:
        -----------
        valor : Any
            Valor a ser formatado
            
        Retorna:
        --------
        Union[str, int, float]: Valor formatado
        """
        if valor is None:
            return ""
        
        # Se for string, retorna como está
        if isinstance(valor, str):
            return valor
        
        # Se for inteiro, retorna como está
        if isinstance(valor, int):
            return valor
        
        # Se for float, aplica formatação específica
        if isinstance(valor, float):
            # Arredonda para 3 casas decimais
            valor_arredondado = round(valor, 3)
            
            # Verifica se o valor arredondado é essencialmente um inteiro
            # Usamos uma tolerância pequena para evitar problemas de ponto flutuante
            if abs(valor_arredondado - round(valor_arredondado)) < 0.0000001:
                return int(round(valor_arredondado))
            
            # Remove zeros desnecessários à direita
            # Converte para string para verificar zeros à direita
            str_valor = f"{valor_arredondado:.10f}"
            str_valor = str_valor.rstrip('0').rstrip('.') if '.' in str_valor else str_valor
            
            # Se após remover zeros ficar vazio, é um inteiro
            if not str_valor:
                return int(valor_arredondado)
            
            # Converte de volta para float, mas mantendo apenas casas decimais necessárias
            return float(str_valor)
        
        # Para outros tipos (bool, etc.), converte para string
        return str(valor)
    
    def _formatar_linha_dados(linha: List) -> List:
        """
        Formata todos os valores de uma linha de dados.
        
        Parâmetros:
        -----------
        linha : List
            Linha de dados a ser formatada
            
        Retorna:
        --------
        List: Linha com valores formatados
        """
        return [_formatar_valor(valor) for valor in linha]
    
    def _formatar_matriz_dados(matriz_dados: List[List]) -> List[List]:
        """
        Formata todos os valores de uma matriz de dados.
        
        Parâmetros:
        -----------
        matriz_dados : List[List]
            Matriz de dados a ser formatada
            
        Retorna:
        --------
        List[List]: Matriz com valores formatados
        """
        return [_formatar_linha_dados(linha) for linha in matriz_dados]
    
    def _extrair_formato_e_validar_caminho():
        """
        Extrai o formato da extensão do arquivo e valida o caminho.
        
        Retorna:
        --------
        str: Formato extraído (xlsx, pdf, txt)
        """
        # Extrai a extensão do arquivo
        _, extensao = os.path.splitext(caminho_arquivo)
        extensao = extensao.lower()
        
        # Verifica se a extensão é suportada
        if extensao not in MAPEAMENTO_EXTENSOES:
            extensoes_suportadas = ', '.join(MAPEAMENTO_EXTENSOES.keys())
            raise ValueError(
                f"Extensão '{extensao}' não suportada. "
                f"Use uma das seguintes extensões: {extensoes_suportadas}"
            )
        
        # Obtém o formato correspondente
        formato = MAPEAMENTO_EXTENSOES[extensao]
        
        # Valida se o diretório existe ou pode ser criado
        diretorio = os.path.dirname(caminho_arquivo)
        if diretorio and not os.path.exists(diretorio):
            try:
                os.makedirs(diretorio, exist_ok=True)
            except (PermissionError, OSError) as e:
                raise ValueError(f"Não foi possível criar o diretório '{diretorio}': {str(e)}")
        
        # Verifica se podemos escrever no diretório
        if diretorio and not os.access(diretorio, os.W_OK):
            raise ValueError(f"Sem permissão de escrita no diretório '{diretorio}'")
        
        return formato
    
    def _validar_parametros(formato: str):
        """Valida os parâmetros de entrada."""
        if not isinstance(matriz, list) or not all(isinstance(linha, list) for linha in matriz):
            raise ValueError("O parâmetro 'matriz' deve ser uma lista de listas.")
        
        if not matriz:
            raise ValueError("A matriz não pode estar vazia.")
        
        # Verificar consistência do número de colunas
        if cabecalho is None:
            num_colunas = len(matriz[0])
            start_idx = 1
        else:
            num_colunas = len(cabecalho)
            start_idx = 0
        
        for i, linha in enumerate(matriz[start_idx:], start=start_idx):
            if len(linha) != num_colunas:
                raise ValueError(f"Linha {i} tem {len(linha)} colunas, mas era esperado {num_colunas}.")
    
    def _preparar_dados():
        """
        Prepara os dados separando cabeçalho e conteúdo.
        
        Retorna:
        --------
        tuple: (cabeçalho_final, dados)
        """
        if cabecalho is None:
            # A primeira linha da matriz é o cabeçalho
            cabecalho_final = matriz[0]
            dados = matriz[1:] if len(matriz) > 1 else []
        else:
            cabecalho_final = cabecalho
            dados = matriz
        
        # Formatar os dados (não formata o cabeçalho)
        dados_formatados = _formatar_matriz_dados(dados)
        
        return cabecalho_final, dados_formatados
    
    def _exportar_xls(cabecalho_final, dados, formato: str):
        """Exporta para formato Excel usando openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "A biblioteca 'openpyxl' é necessária para exportação em Excel. "
                "Instale com: pip install openpyxl"
            )
        
        # Criar workbook e planilha
        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]  # Excel limita a 31 caracteres
        
        # Estilos
        header_font = Font(bold=True, size=12)
        cell_font = Font(size=11)
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Adicionar título
        ws.merge_cells('A1:{0}1'.format(get_column_letter(len(cabecalho_final))))
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Adicionar cabeçalho
        for col_idx, cabecalho_texto in enumerate(cabecalho_final, start=1):
            cell = ws.cell(row=3, column=col_idx, value=cabecalho_texto)
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
        
        # Adicionar dados
        for row_idx, linha in enumerate(dados, start=4):
            for col_idx, valor in enumerate(linha, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = cell_font
                cell.alignment = alignment
                cell.border = thin_border
                
                # Aplica formatação numérica adequada
                if isinstance(valor, (int, float)):
                    # Formato personalizado: mostra apenas casas decimais necessárias
                    cell.number_format = '0.###'  # Até 3 casas decimais
        
        # Ajustar largura das colunas
        for col_idx, cabecalho_texto in enumerate(cabecalho_final, start=1):
            max_length = len(str(cabecalho_texto))
            for linha in dados:
                valor = linha[col_idx-1] if col_idx-1 < len(linha) else ""
                max_length = max(max_length, len(str(valor)))
            
            adjusted_width = min(max_length + 2, 50)  # Limita a largura máxima
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Salvar arquivo
        wb.save(caminho_arquivo)
        return caminho_arquivo
    
    def _exportar_pdf(cabecalho_final, dados, opcoes: Optional[Dict] = None):
        """Exporta para formato PDF usando reportlab."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            raise ImportError(
                "A biblioteca 'reportlab' é necessária para exportação em PDF. "
                "Instale com: pip install reportlab"
            )
        
        # Configurações padrão
        config = {
            'pagesize': landscape(A4),
            'margem': 1.5 * cm,
            'tamanho_fonte_inicial': 10,
            'tamanho_fonte_minimo': 6,
            'cor_cabecalho': colors.HexColor('#2c3e50'),
            'cor_fundo': colors.HexColor('#f8f9fa'),
            'cor_linha_alternada': colors.HexColor('#e9ecef'),
        }
        
        # Atualizar com opções personalizadas
        if opcoes_exportacao:
            config.update(opcoes_exportacao.get('pdf', {}))
        
        # Preparar todos os dados para a tabela (cabeçalho + dados)
        todos_dados = [cabecalho_final] + dados
        
        # Configurações de página
        LARGURA_PAGINA, ALTURA_PAGINA = config['pagesize']
        MARGEM = config['margem']
        
        # Área útil da página
        largura_util = LARGURA_PAGINA - 2 * MARGEM
        altura_util = ALTURA_PAGINA - 2 * MARGEM
        
        # Preparar dados
        num_colunas = len(cabecalho_final)
        
        # Ajustar fonte e larguras
        TAMANHO_FONTE_INICIAL = config['tamanho_fonte_inicial']
        TAMANHO_FONTE_MINIMO = config['tamanho_fonte_minimo']
        tamanho_fonte = TAMANHO_FONTE_INICIAL
        
        def calcular_larguras_colunas(tamanho_fonte_actual):
            """Calcula larguras de coluna baseadas no conteúdo."""
            larguras = []
            
            for col_idx in range(num_colunas):
                # Começar com o cabeçalho
                max_largura = len(str(cabecalho_final[col_idx]))
                
                # Verificar todos os dados
                for linha in todos_dados[1:]:
                    if col_idx < len(linha):
                        conteudo = str(linha[col_idx])
                        largura_conteudo = len(conteudo)
                        max_largura = max(max_largura, largura_conteudo)
                
                # Converter para pontos (estimativa)
                largura_pts = max_largura * tamanho_fonte_actual * 0.6
                largura_pts += tamanho_fonte_actual * 1.2  # Margem interna
                
                larguras.append(largura_pts)
            
            return larguras
        
        def ajustar_larguras_para_pagina(larguras, largura_disponivel):
            """Ajusta as larguras para caber na página."""
            largura_total = sum(larguras)
            
            if largura_total <= largura_disponivel:
                return larguras
            
            # Calcular fator de redução
            fator_reducao = largura_disponivel / largura_total
            larguras_ajustadas = [largura * fator_reducao for largura in larguras]
            
            return larguras_ajustadas
        
        # Ajustar fonte e larguras
        larguras_finais = []
        
        while tamanho_fonte >= TAMANHO_FONTE_MINIMO:
            larguras_colunas = calcular_larguras_colunas(tamanho_fonte)
            larguras_ajustadas = ajustar_larguras_para_pagina(larguras_colunas, largura_util)
            
            # Verificar se as larguras mínimas são aceitáveis
            largura_minima_aceitavel = tamanho_fonte * 2
            if all(largura >= largura_minima_aceitavel for largura in larguras_ajustadas):
                larguras_finais = larguras_ajustadas
                break
            
            # Reduzir fonte e tentar novamente
            tamanho_fonte -= 0.5
        
        # Se não encontrou configuração boa, usar fonte mínima
        if not larguras_finais:
            tamanho_fonte = TAMANHO_FONTE_MINIMO
            larguras_colunas = calcular_larguras_colunas(tamanho_fonte)
            larguras_finais = ajustar_larguras_para_pagina(larguras_colunas, largura_util)
        
        # Criar documento
        doc = SimpleDocTemplate(
            caminho_arquivo,
            pagesize=config['pagesize'],
            leftMargin=MARGEM,
            rightMargin=MARGEM,
            topMargin=MARGEM,
            bottomMargin=MARGEM
        )
        
        # Estilos para parágrafos
        styles = getSampleStyleSheet()
        
        # Estilo para título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=20,
            alignment=1
        )
        
        # Estilo para células
        celula_style = ParagraphStyle(
            'CustomCell',
            parent=styles['Normal'],
            fontSize=tamanho_fonte,
            wordWrap='CJK'
        )
        
        # Estilo para cabeçalho
        cabecalho_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=tamanho_fonte + 1,
            textColor=colors.white,
            alignment=1,
            spaceBefore=5,
            spaceAfter=5
        )
        
        # Criar elementos do documento
        elementos = []
        
        # Adicionar título
        elementos.append(Paragraph(titulo, titulo_style))
        elementos.append(Spacer(1, 0.5 * cm))
        
        # Preparar dados formatados
        dados_formatados = []
        
        # Formatar cabeçalho
        cabecalho_formatado = []
        for cabecalho_texto in cabecalho_final:
            cabecalho_formatado.append(Paragraph(str(cabecalho_texto), cabecalho_style))
        dados_formatados.append(cabecalho_formatado)
        
        # Formatar dados (já estão formatados numericamente, agora convertemos para Paragraph)
        for linha in dados:
            linha_formatada = []
            for valor in linha:
                texto = str(valor) if valor is not None else ""
                if len(texto) > 100:
                    texto = texto[:97] + "..."
                linha_formatada.append(Paragraph(texto, celula_style))
            
            # Garantir que tenha o mesmo número de colunas que o cabeçalho
            while len(linha_formatada) < num_colunas:
                linha_formatada.append(Paragraph("", celula_style))
            
            dados_formatados.append(linha_formatada)
        
        # Calcular altura de linha
        altura_linha = tamanho_fonte * 1.8
        linhas_por_pagina = int((altura_util - 100) / altura_linha)
        linhas_por_pagina = max(10, linhas_por_pagina)
        
        # Dividir dados em páginas
        total_linhas = len(dados_formatados)
        linhas_processadas = 0
        
        while linhas_processadas < total_linhas:
            if linhas_processadas > 0:
                elementos.append(PageBreak())
                elementos.append(Paragraph(f"{titulo} (continuação)", titulo_style))
                elementos.append(Spacer(1, 0.5 * cm))
            
            if linhas_processadas == 0:
                linhas_na_pagina = min(linhas_por_pagina, total_linhas)
                dados_pagina = dados_formatados[:linhas_na_pagina]
                linhas_processadas = linhas_na_pagina
            else:
                linhas_restantes = total_linhas - linhas_processadas
                linhas_na_pagina = min(linhas_por_pagina - 1, linhas_restantes)
                dados_pagina = [dados_formatados[0]] + dados_formatados[linhas_processadas:linhas_processadas + linhas_na_pagina]
                linhas_processadas += linhas_na_pagina
            
            # Criar tabela para esta página
            tabela = Table(
                dados_pagina,
                colWidths=larguras_finais,
                repeatRows=1
            )
            
            # Estilo da tabela
            estilo_tabela = TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), config['cor_cabecalho']),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), tamanho_fonte + 1),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Dados
                ('BACKGROUND', (0, 1), (-1, -1), config['cor_fundo']),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), tamanho_fonte),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                
                # Grades
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#343a40')),
                
                # Preenchimento
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ])
            
            # Adicionar zebrado
            for i in range(1, len(dados_pagina)):
                if i % 2 == 0:
                    estilo_tabela.add('BACKGROUND', (0, i), (-1, i), config['cor_linha_alternada'])
            
            tabela.setStyle(estilo_tabela)
            elementos.append(tabela)
            
            if linhas_processadas < total_linhas:
                elementos.append(Spacer(1, 1 * cm))
        
        # Adicionar rodapé
        rodape_style = ParagraphStyle(
            'CustomFooter',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.gray,
            alignment=1
        )
        
        rodape_texto = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | " \
                      f"Total de registros: {len(dados)} | " \
                      f"Total de colunas: {num_colunas}"
        
        elementos.append(Spacer(1, 1 * cm))
        elementos.append(Paragraph(rodape_texto, rodape_style))
        
        # Construir documento
        doc.build(elementos)
        return caminho_arquivo
    
    def _exportar_txt(cabecalho_final, dados):
        """Exporta para formato TXT com formatação tabular."""
        # Calcular largura de cada coluna
        larguras_colunas = []
        for col_idx in range(len(cabecalho_final)):
            largura_max = len(str(cabecalho_final[col_idx]))
            for linha in dados:
                if col_idx < len(linha):
                    largura_max = max(largura_max, len(str(linha[col_idx])))
            larguras_colunas.append(largura_max + 2)  # +2 para margem
        
        # Desenhar linha horizontal
        def linha_horizontal():
            linha = "+"
            for largura in larguras_colunas:
                linha += "-" * largura + "+"
            return linha + "\n"
        
        # Formatar célula
        def formatar_celula(texto, largura):
            texto_str = str(texto)
            return " " + texto_str.ljust(largura - 1)
        
        # Construir tabela
        with open(caminho_arquivo, 'w', encoding='utf-8') as f:
            # Título
            f.write("=" * 80 + "\n")
            f.write(titulo.center(80) + "\n")
            f.write("=" * 80 + "\n\n")
            
            # Data de geração
            f.write(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
            
            # Cabeçalho da tabela
            f.write(linha_horizontal())
            
            # Linha do cabeçalho
            f.write("|")
            for i, cabecalho_texto in enumerate(cabecalho_final):
                f.write(formatar_celula(cabecalho_texto, larguras_colunas[i]) + "|")
            f.write("\n")
            f.write(linha_horizontal())
            
            # Dados
            if dados:
                for linha in dados:
                    f.write("|")
                    for i, valor in enumerate(linha):
                        if i < len(larguras_colunas):
                            f.write(formatar_celula(valor, larguras_colunas[i]) + "|")
                    f.write("\n")
                f.write(linha_horizontal())
            else:
                f.write("|" + "Nenhum dado disponível".center(sum(larguras_colunas) + len(larguras_colunas) - 1) + "|\n")
                f.write(linha_horizontal())
            
            # Estatísticas
            f.write(f"\nTotal de registros: {len(dados)}\n")
            f.write(f"Total de colunas: {len(cabecalho_final)}\n")
        
        return caminho_arquivo
    
    # Extrair formato do caminho do arquivo
    formato = _extrair_formato_e_validar_caminho()
    
    # Validação inicial dos parâmetros
    _validar_parametros(formato)
    
    # Preparar dados (separar cabeçalho se necessário)
    cabecalho_final, dados = _preparar_dados()
    
    # Exportar conforme o formato detectado
    try:
        if formato in ['xlsx', 'xls']:
            caminho_gerado = _exportar_xls(cabecalho_final, dados, formato)
        
        elif formato == 'pdf':
            caminho_gerado = _exportar_pdf(cabecalho_final, dados, opcoes_exportacao)
        
        elif formato == 'txt':
            caminho_gerado = _exportar_txt(cabecalho_final, dados)
        
        else:
            # Isso não deve acontecer devido à validação anterior
            raise ValueError(f"Formato '{formato}' não implementado.")
        
        print(f"Arquivo gerado com sucesso: {caminho_gerado}")
        return caminho_gerado
        
    except ImportError as e:
        print(f"Erro de importação: {e}")
        print("Instale as bibliotecas necessárias:")
        print("  Excel: pip install openpyxl")
        print("  PDF: pip install reportlab")
        raise
    except Exception as e:
        print(f"Erro ao exportar tabela: {e}")
        raise


def salvar_figura_png(fig, caminho_arquivo, dpi=300, bbox_inches='tight', pad_inches=0.1):
    """
    Salva uma figura matplotlib em formato PNG.
    
    Parâmetros:
    -----------
    fig : matplotlib.figure.Figure
        Objeto da figura a ser salva
    caminho_arquivo : str
        Caminho completo onde a figura será salva (incluindo nome do arquivo e extensão .png)
    dpi : int, opcional
        Resolução da imagem (pontos por polegada), padrão 300
    bbox_inches : str, opcional
        Controla o bounding box, padrão 'tight' para ajustar automaticamente
    pad_inches : float, opcional
        Preenchimento ao redor da figura em polegadas, padrão 0.1
    
    Exemplo:
    --------
    fig = plt.figure()
    # ... seu código de plotagem ...
    salvar_figura_png(fig, 'meu_grafico.png')
    """
    try:
        # Garante que o diretório existe
        diretorio = os.path.dirname(caminho_arquivo)
        if diretorio and not os.path.exists(diretorio):
            os.makedirs(diretorio)
        
        # Salva a figura
        fig.savefig(
            caminho_arquivo,
            dpi=dpi,
            format='png',
            bbox_inches=bbox_inches,
            pad_inches=pad_inches,
            transparent=False
        )
        
        print(f"Figura salva com sucesso em: {caminho_arquivo}")
        
    except Exception as e:
        print(f"Erro ao salvar figura: {str(e)}")
        raise

# Função de teste
def testar_exportacao():
    """Função para testar a exportação em diferentes formatos."""
    
    # Criar dados de exemplo mais complexos
    dados_exemplo = [
        ["ID", "Nome Completo", "E-mail Corporativo", "Telefone", "Cidade", "Cargo", "Salário (R$)", "Data Admissão", "Status"],
        [1, "João Silva Santos", "joao.silva@empresa.com.br", "(11) 99999-9999", "São Paulo - SP", "Analista de Sistemas Sênior", 8500.50, "15/01/2020", "Ativo"],
        [2, "Maria Santos Oliveira", "maria.oliveira@empresa.com.br", "(21) 98888-8888", "Rio de Janeiro - RJ", "Gerente de Projetos", 12500.75, "10/03/2019", "Ativo"],
        [3, "Pedro Costa Pereira", "pedro.pereira@empresa.com.br", "(31) 97777-7777", "Belo Horizonte - MG", "Desenvolvedor Full Stack", 9500.00, "22/07/2021", "Ativo"],
        [4, "Ana Oliveira Mendes", "ana.mendes@empresa.com.br", "(41) 96666-6666", "Curitiba - PR", "Designer UX/UI", 6800.25, "28/02/2022", "Ativo"],
        [5, "Carlos Mendes Lima", "carlos.lima@empresa.com.br", "(51) 95555-5555", "Porto Alegre - RS", "Coordenador de TI", 11500.00, "05/11/2018", "Inativo"],
        [6, "Fernanda Lima Souza", "fernanda.souza@empresa.com.br", "(61) 94444-4444", "Brasília - DF", "Analista de Dados", 7200.80, "18/09/2020", "Ativo"],
        [7, "Ricardo Souza Almeida", "ricardo.almeida@empresa.com.br", "(71) 93333-3333", "Salvador - BA", "Arquiteto de Software", 13200.00, "30/06/2017", "Ativo"],
        [8, "Juliana Almeida Costa", "juliana.costa@empresa.com.br", "(81) 92222-2222", "Recife - PE", "Product Owner", 11800.50, "12/04/2019", "Ativo"],
        [9, "Roberto Costa Silva", "roberto.silva@empresa.com.br", "(91) 91111-1111", "Belém - PA", "Analista de Qualidade", 6200.00, "25/08/2021", "Ativo"],
        [10, "Patrícia Silva Santos", "patricia.santos@empresa.com.br", "(19) 90000-0000", "Campinas - SP", "Scrum Master", 10500.00, "14/12/2020", "Ativo"],
    ]
    
    # Testar com cabeçalho separado
    cabecalho_exemplo = ["ID", "Nome", "E-mail", "Telefone", "Cidade", "Cargo", "Salário", "Admissão", "Status"]
    
    try:
        print("Testando exportação para Excel...")
        arquivo_excel = exportar_tabela(
            matriz=dados_exemplo[1:],
            titulo="Relatório de Funcionários - Empresa XYZ",
            formato="xls",
            cabecalho=cabecalho_exemplo
        )
        print(f"Excel gerado: {arquivo_excel}\n")
        
        print("Testando exportação para PDF...")
        arquivo_pdf = exportar_tabela(
            matriz=dados_exemplo[1:],
            titulo="Relatório de Funcionários - Empresa XYZ",
            formato="pdf",
            cabecalho=cabecalho_exemplo
        )
        print(f"PDF gerado: {arquivo_pdf}\n")
        
        print("Testando exportação para TXT...")
        arquivo_txt = exportar_tabela(
            matriz=dados_exemplo[1:],
            titulo="Relatório de Funcionários - Empresa XYZ",
            formato="txt",
            cabecalho=cabecalho_exemplo
        )
        print(f"TXT gerado: {arquivo_txt}\n")
        
        # Testar com cabeçalho na matriz
        print("Testando com cabeçalho na matriz...")
        arquivo_txt2 = exportar_tabela(
            matriz=dados_exemplo,
            titulo="Funcionários - Cabeçalho na Matriz",
            formato="txt",
            cabecalho=None
        )
        print(f"TXT (cabeçalho na matriz) gerado: {arquivo_txt2}\n")
        
        print("Todos os testes foram concluídos com sucesso!")
        
    except Exception as e:
        print(f"Erro durante os testes: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Executar teste
    testar_exportacao()